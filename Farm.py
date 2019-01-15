import openpyxl
from openpyxl import Workbook
from pprint import pprint
import warnings
from openpyxl.worksheet.datavalidation import DataValidation

warnings.simplefilter("ignore")
wbTemp = openpyxl.load_workbook('Narrative Template Example.xlsx')
wbNar = openpyxl.load_workbook('2018 Narratives Unmerge.xlsx')
wbMatrix = openpyxl.load_workbook('2019 Master ITGC Matrix.xlsx')
warnings.simplefilter("default")

"""
Control Dictionary: key, value = {'SOC 2 ID': 'SOX ID'}

The following characters for SOC2ID and SOXID are not allowed in 2019 Master ITGC Matrix spreadsheet: 
:  \  /  ?  *  [  or  ]

(The above characters are illegal for Excel Tab naming conventions)
"""
def controlDict():
    soc2ID=[]
    soxID=[]
    wbMatrix.active
    ws = wbMatrix["Matrix"]
    for x in ws["B"]:
        soc2ID.append(x.value)
        for y in ws["C"]:
            soxID.append(y.value)
    controlDict = dict(zip(soc2ID,soxID))
    del controlDict['Farm Credit Bank of Texas\nSOC 2 ID']
    del controlDict[None]
    return controlDict

"""
Functions to get Data from ITGC Master Control Matrix 2019
"""
def getShortDesc():
    wbMatrix.active
    ws = wbMatrix["Matrix"]
    controlShort=[]
    for i in ws["D"]:
        controlShort.append(i.value)
    return controlShort[2:70]

def getFreq():
    wbMatrix.active
    ws = wbMatrix["Matrix"]
    controlFreq=[]
    for i in ws["G"]:
        controlFreq.append(i.value)
    return controlFreq[2:70]

    


"""This is broken. I can solve this issue but for the sake 
of time columns A & F in ITGC master can be transposed to
perform an HLOOKUP across all tabs using SOC2ID as primary key
"""
# def getAVM():
#     wbMatrix.active
#     ws = wbMatrix["Matrix"]
#     controlAVM=[]
#     for i in ws["I"]:
#         controlAVM.append(i.value)
#     return controlAVM[2:70]

def getKeyDesc():
    wbMatrix.active
    ws = wbMatrix["Matrix"]
    controlKeyActivity=[]
    for i in ws["E"]:
        controlKeyActivity.append(i.value)
    return controlKeyActivity[2:70]

def getNarInfo(cellValue):
    ws = wbNar.active
    sheetList=[]
    emptList=[]
    for i, sheet in enumerate(wbNar.sheetnames):
        sheetList.append(sheet)
        ws = wbNar[sheetList[i]]
        for cell in ws["A"]:
            if cell.value == cellValue:
                emptList.append(cell.offset(column=1).value)
    return emptList

  

def createTemplateTabs():
    controlNarrative = getNarInfo('Control Narrative')
    precisionNar = getNarInfo('For review controls, describe the precision of review. Are thresholds utilized?')
    howNar = getNarInfo('How is performance of the control evidenced? ')
    controlDictionary = controlDict()
    shortDesc = getShortDesc()
    keyDesc = getKeyDesc()
    freq = getFreq()
    freqDV = DataValidation(type="list", formula1='"As Needed,Annual,Quarterly,Monthly,Weekly,Daily,Other"', allow_blank=True)
    aVM = DataValidation(type="list", formula1='"Automated/Manual,Automated,Manual"', allow_blank=True)
    impactEnv = DataValidation(type="list", formula1='"Physical Security,Windows,Application,Database,Network,Entity Level,iSeries,Other"', allow_blank=True)
    booleanVal = DataValidation(type="list", formula1='"Yes, No"', allow_blank=True)
    tabList=[]
    source = wbTemp.active
    for k, v in controlDictionary.items():
        tabList.append(str(k)+'('+str(v)+')')
        wbTemp.copy_worksheet(source)
    for i, v in enumerate(tabList):
        wbTemp.active = 2 + i
        copyTabs = wbTemp.active
        copyTabs.title = v
        ws = wbTemp[v]
        ws["B5"] = i + 1
        ws["B4"] = v[-6:]
        ws.add_data_validation(freqDV)
        ws.add_data_validation(aVM)
        ws.add_data_validation(impactEnv)
        ws.add_data_validation(booleanVal)
        freqDV.add(ws["B8"])
        impactEnv.add(ws["B10"])
        aVM.add(ws["B12"])
        booleanVal.add(ws["C22"])
        booleanVal.add(ws["D22"])
 
    for i, v in enumerate(tabList):
            wbTemp.active = 2 + i
            copyTabs = wbTemp.active
            copyTabs.title = v
            ws = wbTemp[v]
            try:
                ws["B6"] = shortDesc[i]
                ws["B8"] = freq[i]
                ws["B7"] = keyDesc[i]
                ws["B16"] = controlNarrative[i]
                ws["B17"] = precisionNar[i]
                ws["B18"] = howNar[i]
            except (IndexError):
                pass
           
    wbTemp.save('Narrative Template Example Test.xlsx')

createTemplateTabs()


                
    


   
    







        
   
        
  
